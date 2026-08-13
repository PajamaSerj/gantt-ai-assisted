from dataclasses import dataclass
from io import BytesIO
from zipfile import BadZipFile

from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict

from app.domain.models import normalize_task_name

REQUIRED_COLUMNS = (
    "задача",
    "описание",
    "исполнитель",
    "длительность",
    "предшественники",
)


class ImportIssue(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    code: str
    message: str
    row: int | None = None
    column: str | None = None


@dataclass(frozen=True, slots=True)
class ParsedTaskRow:
    row_number: int
    name: str
    description: str | None
    assignee: str | None
    duration_workdays: int
    predecessor_names: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ParsedWorkbook:
    rows: tuple[ParsedTaskRow, ...]
    issues: tuple[ImportIssue, ...]


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _optional_text(value: object) -> str | None:
    if _is_blank(value):
        return None
    return str(value).strip()


def _duration(value: object) -> int | None:
    if isinstance(value, bool) or _is_blank(value):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        return int(value) if value.is_integer() and value > 0 else None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit() and int(stripped) > 0:
            return int(stripped)
    return None


def _predecessors(value: object) -> tuple[str, ...] | None:
    if _is_blank(value):
        return ()
    if not isinstance(value, str):
        return None
    return tuple(part.strip() for part in value.split(";") if part.strip())


def parse_xlsx(file_name: str, content: bytes) -> ParsedWorkbook:
    issues: list[ImportIssue] = []
    if not file_name.casefold().endswith(".xlsx"):
        return ParsedWorkbook(
            rows=(),
            issues=(
                ImportIssue(
                    code="INVALID_EXTENSION",
                    message="Only .xlsx workbooks are supported",
                ),
            ),
        )

    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True
        )
    except (
        BadZipFile,
        DefusedXmlException,
        EOFError,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as error:
        return ParsedWorkbook(
            rows=(),
            issues=(
                ImportIssue(
                    code="UNREADABLE_WORKBOOK",
                    message=f"Workbook cannot be read: {error}",
                ),
            ),
        )

    try:
        worksheet = workbook.active
        if worksheet is None:
            return ParsedWorkbook(
                rows=(),
                issues=(
                    ImportIssue(
                        code="MISSING_ACTIVE_WORKSHEET",
                        message="Workbook has no active worksheet",
                    ),
                ),
            )

        row_iterator = enumerate(worksheet.iter_rows(values_only=True), start=1)
        header_number: int | None = None
        header_values: tuple[object, ...] = ()
        for row_number, values in row_iterator:
            if not all(_is_blank(value) for value in values):
                header_number = row_number
                header_values = values
                break

        if header_number is None:
            return ParsedWorkbook(
                rows=(),
                issues=(
                    ImportIssue(
                        code="EMPTY_WORKSHEET",
                        message="Active worksheet is empty",
                    ),
                ),
            )

        normalized_headers = [_normalize_header(value) for value in header_values]
        header_positions: dict[str, int] = {}
        for position, header in enumerate(normalized_headers):
            if header not in REQUIRED_COLUMNS:
                continue
            if header in header_positions:
                issues.append(
                    ImportIssue(
                        code="DUPLICATE_COLUMN",
                        message=f"Column '{header}' appears more than once",
                        row=header_number,
                        column=header,
                    )
                )
            else:
                header_positions[header] = position

        for required in REQUIRED_COLUMNS:
            if required not in header_positions:
                issues.append(
                    ImportIssue(
                        code="MISSING_COLUMN",
                        message=f"Required column '{required}' is missing",
                        row=header_number,
                        column=required,
                    )
                )
        if issues:
            return ParsedWorkbook(rows=(), issues=tuple(issues))

        parsed_rows: list[ParsedTaskRow] = []
        name_rows: dict[str, int] = {}
        for row_number, values in row_iterator:
            if all(_is_blank(value) for value in values):
                continue

            def value_for(column: str) -> object:
                position = header_positions[column]
                return values[position] if position < len(values) else None

            raw_name = value_for("задача")
            name = str(raw_name).strip() if not _is_blank(raw_name) else ""
            duration = _duration(value_for("длительность"))
            predecessor_names = _predecessors(value_for("предшественники"))
            row_has_error = False

            if not name:
                issues.append(
                    ImportIssue(
                        code="MISSING_TASK_NAME",
                        message="Task name is required",
                        row=row_number,
                        column="задача",
                    )
                )
                row_has_error = True
            else:
                try:
                    name = normalize_task_name(name)
                except ValueError as error:
                    issues.append(
                        ImportIssue(
                            code="INVALID_TASK_NAME",
                            message=str(error),
                            row=row_number,
                            column="задача",
                        )
                    )
                    row_has_error = True
                else:
                    normalized_name = name.casefold()
                    if normalized_name in name_rows:
                        issues.append(
                            ImportIssue(
                                code="DUPLICATE_TASK_NAME",
                                message=(
                                    f"Task name '{name}' duplicates row "
                                    f"{name_rows[normalized_name]}"
                                ),
                                row=row_number,
                                column="задача",
                            )
                        )
                        row_has_error = True
                    else:
                        name_rows[normalized_name] = row_number

            if duration is None:
                issues.append(
                    ImportIssue(
                        code="INVALID_DURATION",
                        message="Duration must be a positive integer",
                        row=row_number,
                        column="длительность",
                    )
                )
                row_has_error = True

            if predecessor_names is None:
                issues.append(
                    ImportIssue(
                        code="INVALID_PREDECESSORS",
                        message="Predecessors must be names separated by ';'",
                        row=row_number,
                        column="предшественники",
                    )
                )
                row_has_error = True

            if row_has_error:
                continue

            assert duration is not None
            assert predecessor_names is not None
            parsed_rows.append(
                ParsedTaskRow(
                    row_number=row_number,
                    name=name,
                    description=_optional_text(value_for("описание")),
                    assignee=_optional_text(value_for("исполнитель")),
                    duration_workdays=duration,
                    predecessor_names=predecessor_names,
                )
            )

        if not parsed_rows and not issues:
            issues.append(
                ImportIssue(
                    code="NO_TASK_ROWS",
                    message="Active worksheet contains no task rows",
                )
            )
        return ParsedWorkbook(rows=tuple(parsed_rows), issues=tuple(issues))
    finally:
        workbook.close()
