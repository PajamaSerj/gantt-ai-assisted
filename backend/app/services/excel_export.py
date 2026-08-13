from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.domain.graph import task_index
from app.domain.models import PlanState
from app.domain.validation import validate_plan_schedule

EXPORT_COLUMNS = (
    "ID",
    "задача",
    "описание",
    "исполнитель",
    "длительность",
    "дата начала",
    "дата окончания",
    "предшественники",
)


def export_plan_xlsx(plan: PlanState) -> bytes:
    validate_plan_schedule(plan)
    indexed = task_index(plan.tasks)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plan"
    worksheet.append(EXPORT_COLUMNS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for task in plan.tasks:
        predecessor_names = "; ".join(
            indexed[predecessor_id].name
            for predecessor_id in task.predecessor_ids
        )
        worksheet.append(
            (
                task.public_id,
                task.name,
                task.description,
                task.assignee,
                task.duration_workdays,
                task.start_date,
                task.end_date,
                predecessor_names,
            )
        )
        worksheet.cell(worksheet.max_row, 6).number_format = "yyyy-mm-dd"
        worksheet.cell(worksheet.max_row, 7).number_format = "yyyy-mm-dd"

    widths = (14, 28, 42, 24, 15, 18, 18, 38)
    for position, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + position)].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
