from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

import app.services.excel_import as excel_import_service
from app.domain.changesets import (
    ChangeSet,
    ChangeConflict,
    ChangeSetStatus,
    apply_changeset,
    plan_digest,
)
from app.domain.errors import InvalidChangeSetError
from app.domain.models import PlanState
from app.seed.data import get_seed_plan
from app.services.excel_import import REQUIRED_COLUMNS, parse_xlsx
from app.services.import_planning import ImportMode, prepare_import


def workbook_bytes(
    rows: list[tuple],
    *,
    headers: tuple[str, ...] = REQUIRED_COLUMNS,
    active_second_sheet: bool = False,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    if active_second_sheet:
        worksheet.append(("not", "the", "required", "structure"))
        worksheet = workbook.create_sheet("Active Plan")
        workbook.active = 1
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def prepare_replace(rows: list[tuple], start: date = date(2026, 8, 17)):
    return prepare_import(
        file_name="tasks.xlsx",
        content=workbook_bytes(rows),
        mode=ImportMode.REPLACE,
        date_constraint=start,
        current_plan=PlanState(),
    )


def test_valid_xlsx_normalizes_headers_and_ignores_extra_columns() -> None:
    content = workbook_bytes(
        [("Discovery", "Scope", "Anna", 2, None, "ignored")],
        headers=(
            " ЗАДАЧА ",
            "Описание",
            "ИСПОЛНИТЕЛЬ",
            " длительность ",
            "Предшественники",
            "Extra",
        ),
    )

    parsed = parse_xlsx("TASKS.XLSX", content)

    assert parsed.issues == ()
    assert parsed.rows[0].name == "Discovery"
    assert parsed.rows[0].duration_workdays == 2


def test_duplicate_unknown_extra_columns_are_ignored() -> None:
    content = workbook_bytes(
        [("Discovery", "Scope", "Anna", 2, None, "first", "second")],
        headers=(*REQUIRED_COLUMNS, "note", " NOTE "),
    )

    parsed = parse_xlsx("tasks.xlsx", content)

    assert parsed.issues == ()
    assert [row.name for row in parsed.rows] == ["Discovery"]


def test_duplicate_required_column_is_rejected() -> None:
    content = workbook_bytes(
        [("Discovery", "Scope", "Anna", 2, None, "Duplicate name")],
        headers=(*REQUIRED_COLUMNS, " ЗАДАЧА "),
    )

    parsed = parse_xlsx("tasks.xlsx", content)

    assert [(issue.code, issue.column) for issue in parsed.issues] == [
        ("DUPLICATE_COLUMN", "задача")
    ]
    assert parsed.issues[0].row == 1
    assert parsed.issues[0].message == (
        "Обязательная колонка «задача» указана больше одного раза."
    )


def test_only_active_worksheet_is_processed() -> None:
    content = workbook_bytes(
        [("Discovery", None, None, 1, None)], active_second_sheet=True
    )

    parsed = parse_xlsx("tasks.xlsx", content)

    assert parsed.issues == ()
    assert [row.name for row in parsed.rows] == ["Discovery"]


def test_non_xlsx_file_is_rejected() -> None:
    parsed = parse_xlsx("tasks.xls", b"not a workbook")

    assert [issue.code for issue in parsed.issues] == ["INVALID_EXTENSION"]
    assert parsed.issues[0].message == "Поддерживаются только файлы .xlsx."


def test_unreadable_xlsx_is_rejected() -> None:
    parsed = parse_xlsx("tasks.xlsx", b"not a zip archive")

    assert [issue.code for issue in parsed.issues] == ["UNREADABLE_WORKBOOK"]
    assert parsed.issues[0].message == (
        "Не удалось прочитать Excel-файл. Проверьте, что файл не повреждён "
        "и имеет формат .xlsx."
    )
    assert "zip" not in parsed.issues[0].message.casefold()


def test_missing_active_worksheet_message_is_russian(monkeypatch) -> None:
    class WorkbookWithoutActiveSheet:
        active = None

        def close(self) -> None:
            pass

    monkeypatch.setattr(
        excel_import_service,
        "load_workbook",
        lambda *_args, **_kwargs: WorkbookWithoutActiveSheet(),
    )

    parsed = parse_xlsx("tasks.xlsx", b"valid enough for the stub")

    assert [(issue.code, issue.row, issue.column) for issue in parsed.issues] == [
        ("MISSING_ACTIVE_WORKSHEET", None, None)
    ]
    assert parsed.issues[0].message == "В книге нет активного листа."


def test_empty_and_header_only_workbooks_have_distinct_russian_messages() -> None:
    empty = Workbook()
    empty_output = BytesIO()
    empty.save(empty_output)
    empty.close()

    empty_parsed = parse_xlsx("empty.xlsx", empty_output.getvalue())
    header_only = parse_xlsx("headers.xlsx", workbook_bytes([]))

    assert [(issue.code, issue.message) for issue in empty_parsed.issues] == [
        ("EMPTY_WORKSHEET", "Активный лист пуст.")
    ]
    assert [(issue.code, issue.message) for issue in header_only.issues] == [
        ("NO_TASK_ROWS", "На активном листе нет строк с задачами.")
    ]


def test_missing_columns_are_reported_together() -> None:
    content = workbook_bytes([], headers=("задача", "длительность"))

    parsed = parse_xlsx("tasks.xlsx", content)

    assert {issue.column for issue in parsed.issues} == {
        "описание",
        "исполнитель",
        "предшественники",
    }
    assert {issue.message for issue in parsed.issues} == {
        "Не найдена обязательная колонка «описание».",
        "Не найдена обязательная колонка «исполнитель».",
        "Не найдена обязательная колонка «предшественники».",
    }
    assert {issue.row for issue in parsed.issues} == {1}


def test_invalid_durations_are_collected_with_row_numbers() -> None:
    content = workbook_bytes(
        [
            ("A", None, None, 0, None),
            ("B", None, None, 1.5, None),
            ("C", None, None, "two", None),
        ]
    )

    parsed = parse_xlsx("tasks.xlsx", content)

    assert [(issue.code, issue.row) for issue in parsed.issues] == [
        ("INVALID_DURATION", 2),
        ("INVALID_DURATION", 3),
        ("INVALID_DURATION", 4),
    ]
    assert {issue.message for issue in parsed.issues} == {
        "Длительность должна быть положительным целым числом рабочих дней."
    }
    assert {issue.column for issue in parsed.issues} == {"длительность"}


def test_row_field_errors_are_russian_and_keep_metadata() -> None:
    parsed = parse_xlsx(
        "tasks.xlsx",
        workbook_bytes([(None, None, None, "two", 42)]),
    )

    assert [
        (issue.code, issue.message, issue.row, issue.column)
        for issue in parsed.issues
    ] == [
        ("MISSING_TASK_NAME", "Укажите название задачи.", 2, "задача"),
        (
            "INVALID_DURATION",
            "Длительность должна быть положительным целым числом рабочих дней.",
            2,
            "длительность",
        ),
        (
            "INVALID_PREDECESSORS",
            "Предшественники должны быть указаны названиями задач через «;».",
            2,
            "предшественники",
        ),
    ]


def test_duplicate_task_names_are_rejected_case_insensitively() -> None:
    preparation = prepare_replace(
        [
            ("Backend", None, None, 1, None),
            ("backend", None, None, 1, None),
        ]
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.issues[0].code == "DUPLICATE_TASK_NAME"
    assert preparation.issues[0].message == (
        "Название задачи «backend» повторяет строку 2. "
        "Названия задач должны быть уникальными."
    )
    assert preparation.issues[0].row == 3
    assert preparation.issues[0].column == "задача"


def test_task_name_with_reserved_separator_is_rejected_atomically() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="append.xlsx",
        content=workbook_bytes(
            [
                ("Valid task", None, None, 1, None),
                ("Invalid; task", None, None, 1, None),
            ]
        ),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.changeset is None
    assert preparation.unchanged_plan == current
    assert [(issue.code, issue.row, issue.column) for issue in preparation.issues] == [
        ("INVALID_TASK_NAME", 3, "задача")
    ]
    assert preparation.issues[0].message == (
        "Название задачи не может содержать «;»: этот символ используется "
        "как разделитель предшественников в Excel."
    )


def test_unknown_predecessor_is_rejected_atomically() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="tasks.xlsx",
        content=workbook_bytes(
            [("Review", None, None, 1, "Missing task")]
        ),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.changeset is None
    assert preparation.unchanged_plan == current
    assert preparation.issues[0].code == "UNKNOWN_PREDECESSOR"
    assert preparation.issues[0].message == (
        "Предшественник «Missing task» не найден ни в загружаемом Excel, "
        "ни в текущем плане."
    )
    assert preparation.issues[0].row == 2
    assert preparation.issues[0].column == "предшественники"


def test_replace_unknown_predecessors_explain_replace_scope_in_russian() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="append-oriented.xlsx",
        content=workbook_bytes(
            [
                (
                    "Публикация релиза",
                    None,
                    None,
                    1,
                    "Интеграция приложения; Сквозное тестирование",
                )
            ]
        ),
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.changeset is None
    assert preparation.unchanged_plan == current
    assert [
        (issue.code, issue.message, issue.row, issue.column)
        for issue in preparation.issues
    ] == [
        (
            "UNKNOWN_PREDECESSOR",
            "Предшественник «Интеграция приложения» не найден. В режиме замены "
            "он должен быть отдельной задачей в загружаемом Excel.",
            2,
            "предшественники",
        ),
        (
            "UNKNOWN_PREDECESSOR",
            "Предшественник «Сквозное тестирование» не найден. В режиме замены "
            "он должен быть отдельной задачей в загружаемом Excel.",
            2,
            "предшественники",
        ),
    ]


def test_self_reference_is_rejected() -> None:
    preparation = prepare_replace(
        [("Backend", None, None, 1, "Backend")]
    )

    assert (
        preparation.issues[0].code,
        preparation.issues[0].message,
        preparation.issues[0].row,
        preparation.issues[0].column,
    ) == (
        "SELF_REFERENCE",
        "Задача «Backend» не может зависеть сама от себя.",
        2,
        "предшественники",
    )


def test_duplicate_predecessor_is_rejected_in_russian() -> None:
    preparation = prepare_replace(
        [
            ("Backend", None, None, 1, None),
            ("Review", None, None, 1, "Backend; Backend"),
        ]
    )

    assert [
        (issue.code, issue.message, issue.row, issue.column)
        for issue in preparation.issues
    ] == [
        (
            "DUPLICATE_PREDECESSOR",
            "Предшественник «Backend» указан повторно.",
            3,
            "предшественники",
        )
    ]


def test_dependency_cycle_is_rejected() -> None:
    preparation = prepare_replace(
        [
            ("A", None, None, 1, "C"),
            ("B", None, None, 1, "A"),
            ("C", None, None, 1, "B"),
        ]
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.issues[0].code == "DEPENDENCY_CYCLE"
    assert preparation.issues[0].message == (
        "Обнаружен цикл зависимостей: "
        "TASK-001 → TASK-003 → TASK-002 → TASK-001."
    )
    assert preparation.issues[0].column == "предшественники"


def test_invalid_current_plan_error_is_localized_at_import_boundary() -> None:
    current = get_seed_plan()
    broken_task = current.tasks[0].model_copy(
        update={"start_date": date(2026, 1, 31), "end_date": date(2026, 2, 2)}
    )
    broken_plan = current.model_copy(
        update={"tasks": (broken_task, *current.tasks[1:])}
    )

    preparation = prepare_import(
        file_name="append.xlsx",
        content=workbook_bytes([("Новая задача", None, None, 1, None)]),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 3),
        current_plan=broken_plan,
    )

    assert [(issue.code, issue.message) for issue in preparation.issues] == [
        (
            "INVALID_CURRENT_PLAN",
            "Текущий план содержит некорректное расписание для TASK-001.",
        )
    ]


def test_import_changeset_conflict_does_not_expose_english_domain_message(
    monkeypatch,
) -> None:
    current = get_seed_plan()

    def invalid_changeset(source_plan, requested_changes):
        return ChangeSet(
            source_plan_digest=plan_digest(source_plan),
            requested_changes=tuple(requested_changes),
            conflicts=(
                ChangeConflict(
                    code="ScheduleValidationError",
                    message="TASK-001 violates Finish-to-Start constraints",
                    task_public_id="TASK-001",
                ),
            ),
            status=ChangeSetStatus.INVALID,
        )

    monkeypatch.setattr(
        "app.services.import_planning.prepare_changeset",
        invalid_changeset,
    )

    preparation = prepare_import(
        file_name="replace.xlsx",
        content=workbook_bytes([("Новая задача", None, None, 1, None)]),
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    assert [(issue.code, issue.message) for issue in preparation.issues] == [
        (
            "ScheduleValidationError",
            "Расписание задачи TASK-001 нарушает правила планирования.",
        )
    ]


def test_replace_generates_ids_and_consolidated_dependency_impacts() -> None:
    preparation = prepare_replace(
        [
            ("A", None, "Anna", 2, None),
            ("B", None, "Boris", 1, "A"),
            ("C", None, "Clara", 1, "B"),
        ]
    )

    assert preparation.status == ChangeSetStatus.CONFIRMATION_REQUIRED.value
    changeset = preparation.changeset
    assert [task.public_id for task in changeset.proposed_plan.tasks] == [
        "TASK-001",
        "TASK-002",
        "TASK-003",
    ]
    assert [impact.public_id for impact in changeset.proposed_impacts] == [
        "TASK-002",
        "TASK-003",
    ]


def test_blank_rows_do_not_create_task_id_gaps() -> None:
    preparation = prepare_import(
        file_name="tasks.xlsx",
        content=workbook_bytes(
            [
                ("A", None, None, 1, None),
                (None, None, None, None, None),
                ("B", None, None, 1, None),
            ]
        ),
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 8, 17),
        current_plan=PlanState(),
    )

    assert preparation.status == ChangeSetStatus.AUTO_APPLICABLE.value
    assert [task.public_id for task in preparation.changeset.proposed_plan.tasks] == [
        "TASK-001",
        "TASK-002",
    ]


def test_append_resolves_existing_task_and_continues_public_ids() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="append.xlsx",
        content=workbook_bytes(
            [("Retrospective", None, "Anna", 1, "Подготовка демо")]
        ),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 2),
        current_plan=current,
    )

    assert preparation.status == ChangeSetStatus.CONFIRMATION_REQUIRED.value
    appended = preparation.changeset.proposed_plan.tasks[-1]
    assert appended.public_id == "TASK-008"
    assert appended.predecessor_ids == (current.tasks[-1].internal_id,)
    assert appended.start_date == date(2026, 3, 3)


def test_append_resolves_dependencies_between_incoming_tasks() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="append.xlsx",
        content=workbook_bytes(
            [
                ("Release notes", None, "Anna", 1, None),
                ("Publish notes", None, "Anna", 1, "Release notes"),
            ]
        ),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    new_tasks = preparation.changeset.proposed_plan.tasks[-2:]
    assert [task.public_id for task in new_tasks] == ["TASK-008", "TASK-009"]
    assert new_tasks[1].predecessor_ids == (new_tasks[0].internal_id,)
    assert [task.public_id for task in current.tasks] == [
        f"TASK-{number:03d}" for number in range(1, 8)
    ]


def test_append_duplicate_across_current_and_incoming_is_atomic() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="append.xlsx",
        content=workbook_bytes(
            [(current.tasks[0].name, None, None, 1, None)]
        ),
        mode=ImportMode.APPEND,
        date_constraint=date(2026, 3, 3),
        current_plan=current,
    )

    assert preparation.status == "VALIDATION_FAILED"
    assert preparation.changeset is None
    assert preparation.unchanged_plan == current
    assert preparation.issues[0].code == "DUPLICATE_TASK_NAME"
    assert preparation.issues[0].message == (
        f"Задача «{current.tasks[0].name}» уже существует в текущем плане "
        f"как {current.tasks[0].public_id}."
    )
    assert preparation.issues[0].row == 2
    assert preparation.issues[0].column == "задача"


def test_weekend_import_date_is_normalized_with_preview() -> None:
    preparation = prepare_replace(
        [("A", None, None, 1, None)], start=date(2026, 8, 22)
    )

    assert preparation.status == ChangeSetStatus.CONFIRMATION_REQUIRED.value
    normalization = preparation.changeset.date_normalizations[0]
    assert normalization.requested_date == date(2026, 8, 22)
    assert normalization.normalized_date == date(2026, 8, 24)
    assert preparation.changeset.proposed_plan.tasks[0].start_date == date(
        2026, 8, 24
    )


def test_import_with_identical_final_plan_returns_no_change(
    monkeypatch,
) -> None:
    current = get_seed_plan()

    def identical_changeset(source_plan, requested_changes):
        return ChangeSet(
            source_plan_digest=plan_digest(source_plan),
            requested_changes=tuple(requested_changes),
            status=ChangeSetStatus.CONFIRMATION_REQUIRED,
            proposed_plan=source_plan,
        )

    monkeypatch.setattr(
        "app.services.import_planning.prepare_changeset",
        identical_changeset,
    )

    preparation = prepare_import(
        file_name="tasks.xlsx",
        content=workbook_bytes([("A", None, None, 1, None)]),
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 8, 17),
        current_plan=current,
    )

    assert preparation.status == "NO_CHANGE"
    assert preparation.unchanged_plan == current
    assert preparation.changeset is None
    assert preparation.issues == ()


def test_replace_changeset_rejects_changed_current_snapshot() -> None:
    current = get_seed_plan()
    preparation = prepare_import(
        file_name="tasks.xlsx",
        content=workbook_bytes([("A", None, None, 1, None)]),
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 8, 17),
        current_plan=current,
    )
    changed = PlanState(
        tasks=(
            current.tasks[0].model_copy(update={"assignee": "Changed"}),
            *current.tasks[1:],
        )
    )

    with pytest.raises(InvalidChangeSetError, match="differs"):
        apply_changeset(changed, preparation.changeset, confirmed=True)
