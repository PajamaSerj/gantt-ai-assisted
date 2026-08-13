import asyncio
import json
from io import BytesIO

import httpx
from openpyxl import Workbook, load_workbook

from app.main import app
from app.seed.data import get_seed_plan


async def request(path: str) -> httpx.Response:
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(
        transport=transport, base_url="http://testserver"
    ) as client:
        return await client.get(path)


async def post(
    path: str,
    *,
    json_body: dict | None = None,
    data: dict | None = None,
    files: dict | None = None,
) -> httpx.Response:
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(
        transport=transport, base_url="http://testserver"
    ) as client:
        return await client.post(path, json=json_body, data=data, files=files)


def workbook_bytes(rows: list[tuple]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        ("задача", "описание", "исполнитель", "длительность", "предшественники")
    )
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_health_endpoint() -> None:
    response = asyncio.run(request("/api/health"))

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_seed_endpoint_returns_fixed_plan_state() -> None:
    response = asyncio.run(request("/api/seed"))

    assert response.status_code == 200
    body = response.json()
    assert list(body) == ["tasks"]
    assert len(body["tasks"]) == 7
    assert body["tasks"][0]["public_id"] == "TASK-001"
    assert body["tasks"][0]["start_date"] == "2026-02-02"


def test_import_and_guarded_apply_endpoints() -> None:
    source_plan = {"tasks": []}
    import_response = asyncio.run(
        post(
            "/api/import",
            data={
                "mode": "replace",
                "date_constraint": "2026-08-17",
                "current_plan": json.dumps(source_plan),
            },
            files={
                "file": (
                    "tasks.xlsx",
                    workbook_bytes(
                        [
                            ("A", None, None, 1, None),
                            ("B", None, None, 1, "A"),
                        ]
                    ),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    )

    assert import_response.status_code == 200
    prepared = import_response.json()
    assert prepared["status"] == "CONFIRMATION_REQUIRED"
    assert prepared["unchanged_plan"] == source_plan

    apply_response = asyncio.run(
        post(
            "/api/changesets/apply",
            json_body={
                "current_plan": source_plan,
                "changeset": prepared["changeset"],
                "choice": "apply_all",
            },
        )
    )

    assert apply_response.status_code == 200
    applied = apply_response.json()
    assert applied["status"] == "applied"
    assert [task["public_id"] for task in applied["plan"]["tasks"]] == [
        "TASK-001",
        "TASK-002",
    ]


def test_invalid_import_keeps_current_plan_unchanged() -> None:
    source_plan = get_seed_plan().model_dump(mode="json")
    response = asyncio.run(
        post(
            "/api/import",
            data={
                "mode": "append",
                "date_constraint": "2026-08-17",
                "current_plan": json.dumps(source_plan),
            },
            files={
                "file": (
                    "tasks.xlsx",
                    workbook_bytes([("Broken", None, None, 0, None)]),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "VALIDATION_FAILED"
    assert body["changeset"] is None
    assert body["unchanged_plan"] == source_plan
    assert body["errors"][0]["code"] == "INVALID_DURATION"


def test_cancel_endpoint_returns_unchanged_plan() -> None:
    source_plan = {"tasks": []}
    import_response = asyncio.run(
        post(
            "/api/import",
            data={
                "mode": "replace",
                "date_constraint": "2026-08-17",
                "current_plan": json.dumps(source_plan),
            },
            files={
                "file": (
                    "tasks.xlsx",
                    workbook_bytes([("A", None, None, 1, None)]),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    ).json()

    response = asyncio.run(
        post(
            "/api/changesets/apply",
            json_body={
                "current_plan": source_plan,
                "changeset": import_response["changeset"],
                "choice": "cancel",
            },
        )
    )

    assert response.status_code == 200
    assert response.json() == {"status": "cancelled", "plan": source_plan}


def test_export_endpoint_returns_openable_xlsx() -> None:
    plan = get_seed_plan().model_dump(mode="json")
    response = asyncio.run(post("/api/export", json_body=plan))

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.active.cell(2, 1).value == "TASK-001"
    workbook.close()
