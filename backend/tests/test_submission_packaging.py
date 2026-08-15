import asyncio
import json
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app.main import app


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_WORKBOOK = REPOSITORY_ROOT / "sample" / "sample_tasks.xlsx"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


async def post(
    path: str,
    *,
    json_body: dict | None = None,
    data: dict | None = None,
    files: dict | None = None,
) -> httpx.Response:
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(
        transport=transport,
        base_url="http://testserver",
    ) as client:
        return await client.post(path, json=json_body, data=data, files=files)


def test_committed_sample_workbook_passes_real_replace_import() -> None:
    content = SAMPLE_WORKBOOK.read_bytes()
    source_plan = {"tasks": []}

    workbook = load_workbook(SAMPLE_WORKBOOK, read_only=False, data_only=False)
    try:
        worksheet = workbook.active
        assert worksheet.title == "Задачи"
        assert worksheet.max_row == 9
        assert worksheet.max_column == 5
        assert list(worksheet.values)[0] == (
            "задача",
            "описание",
            "исполнитель",
            "длительность",
            "предшественники",
        )
        assert "SampleTasksTable" in worksheet.tables
        assert all(
            isinstance(worksheet.cell(row=row, column=4).value, int)
            for row in range(2, 10)
        )
        assert not any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in worksheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()

    response = asyncio.run(
        post(
            "/api/import",
            data={
                "mode": "replace",
                "date_constraint": "2026-09-07",
                "current_plan": json.dumps(source_plan),
            },
            files={
                "file": (SAMPLE_WORKBOOK.name, content, XLSX_MEDIA_TYPE),
            },
        )
    )

    assert response.status_code == 200
    prepared = response.json()
    assert prepared["status"] == "CONFIRMATION_REQUIRED"
    assert prepared["unchanged_plan"] == source_plan
    assert prepared["errors"] == []
    assert prepared["changeset"] is not None

    proposed_tasks = prepared["changeset"]["proposed_plan"]["tasks"]
    assert [task["public_id"] for task in proposed_tasks] == [
        f"TASK-{number:03d}" for number in range(1, 9)
    ]
    assert [task["name"] for task in proposed_tasks] == [
        "Исследование пользователей",
        "Проектирование архитектуры",
        "Прототип API",
        "Концепция интерфейса",
        "Реализация frontend",
        "Интеграция приложения",
        "Приёмочное тестирование",
        "Подготовка запуска",
    ]
    assert proposed_tasks[0]["start_date"] == "2026-09-07"
    assert proposed_tasks[1]["start_date"] == "2026-09-09"
    assert proposed_tasks[3]["start_date"] == "2026-09-09"
    assert len(proposed_tasks[5]["predecessor_ids"]) == 2
    assert {task["assignee"] for task in proposed_tasks} == {
        "Анна",
        "Илья",
        "Мария",
        "Олег",
        "Сергей",
    }

    apply_response = asyncio.run(
        post(
            "/api/changesets/apply",
            json_body={
                "current_plan": source_plan,
                "changeset": prepared["changeset"],
                "choice": "apply_all",
            },
        )
    )

    assert apply_response.status_code == 200
    applied = apply_response.json()
    assert applied["status"] == "applied"
    assert applied["plan"] == prepared["changeset"]["proposed_plan"]
