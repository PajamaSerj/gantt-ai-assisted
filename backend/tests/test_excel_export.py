from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.domain.changesets import ChangeSetStatus
from app.domain.models import PlanState
from app.seed.data import get_seed_plan
from app.services.excel_export import EXPORT_COLUMNS, export_plan_xlsx
from app.services.import_planning import ImportMode, prepare_import


def test_export_contains_human_readable_ids_dates_and_predecessors() -> None:
    plan = get_seed_plan()
    content = export_plan_xlsx(plan)
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active

    assert tuple(cell.value for cell in worksheet[1]) == EXPORT_COLUMNS
    assert worksheet.cell(2, 1).value == "TASK-001"
    assert worksheet.cell(3, 8).value == plan.tasks[0].name
    assert worksheet.cell(2, 6).value.date() == plan.tasks[0].start_date
    assert "00000000" not in str(tuple(worksheet.values))
    workbook.close()


def test_export_can_be_reimported_through_normal_mvp_contract() -> None:
    original = get_seed_plan()
    content = export_plan_xlsx(original)

    preparation = prepare_import(
        file_name="exported-plan.xlsx",
        content=content,
        mode=ImportMode.REPLACE,
        date_constraint=date(2026, 2, 2),
        current_plan=PlanState(),
    )

    assert preparation.status == ChangeSetStatus.CONFIRMATION_REQUIRED.value
    reimported = preparation.changeset.proposed_plan
    assert [task.name for task in reimported.tasks] == [
        task.name for task in original.tasks
    ]
    assert [task.public_id for task in reimported.tasks] == [
        f"TASK-{number:03d}" for number in range(1, len(original.tasks) + 1)
    ]
    assert [task.internal_id for task in reimported.tasks] != [
        task.internal_id for task in original.tasks
    ]
